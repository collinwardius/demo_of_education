#!/usr/bin/env python3
"""
Historical Education Data Cleaner and Validator
Post-processes Textract output to fix common OCR errors and validate data quality
"""

import pandas as pd
import numpy as np
import re
import json
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import logging
from datetime import datetime

class HistoricalDataCleaner:
    def __init__(self):
        """Initialize the data cleaner with common patterns and validation rules."""
        self.setup_logging()
        
        # Common OCR error patterns for numbers
        self.number_fixes = {
            r'(\d+)\.(\d{3})': r'\1,\2',  # 10.066 -> 10,066
            r'(\d+)\s+(\d{3})': r'\1,\2',  # 10 066 -> 10,066
            r'(\d+)O(\d+)': r'\1\2',      # 1O5 -> 105 (O instead of 0)
            r'(\d+)l(\d+)': r'\1\2',      # 1l5 -> 115 (l instead of 1)
            r'S(\d+)': r'5\1',            # S12 -> 512 (S instead of 5)
            r'(\d+)S': r'\g<1>5',         # 12S -> 125
        }
        
        # Common text fixes for education data
        self.text_fixes = {
            'Baccalau-': 'Baccalaureate',
            'Undergrad-': 'Undergraduate',
            'ses-': 'session',
            'pro-': 'professional',
            'engineer-': 'engineering',
            'fessional': 'professional',
            'uate': 'ate',
        }
        
        # Expected column patterns for validation
        self.education_columns = {
            'state': {'pattern': r'^[A-Z][a-z]', 'type': 'categorical'},
            'institutions': {'pattern': r'^\d+$', 'type': 'integer', 'range': (1, 1000)},
            'faculty': {'pattern': r'^[\d,]+$', 'type': 'integer', 'range': (0, 100000)},
            'enrollment': {'pattern': r'^[\d,]+$', 'type': 'integer', 'range': (0, 1000000)},
            'degrees': {'pattern': r'^[\d,]+$', 'type': 'integer', 'range': (0, 50000)},
        }
        
        # Validation flags
        self.validation_flags = []
    
    def setup_logging(self):
        """Setup logging for data cleaning process."""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('data_cleaning.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def clean_dataframe(self, df: pd.DataFrame, document_source: str = "", 
                        include_metadata: bool = True) -> pd.DataFrame:
        """
        Clean and validate a DataFrame extracted from Textract.
        
        Args:
            df: Raw DataFrame from Textract
            document_source: Source document name for tracking
            include_metadata: Whether to include validation metadata columns
            
        Returns:
            Cleaned DataFrame with optional validation flags
        """
        self.logger.info(f"Starting data cleaning for {document_source}")
        
        # Make a copy to avoid modifying original
        cleaned_df = df.copy()
        
        # Step 1: Clean headers
        cleaned_df = self._clean_headers(cleaned_df)
        
        # Step 2: Clean numeric data
        cleaned_df = self._clean_numeric_columns(cleaned_df)
        
        # Step 3: Clean text data
        cleaned_df = self._clean_text_columns(cleaned_df)
        
        # Step 4: Validate data quality
        cleaned_df = self._validate_data(cleaned_df, document_source)
        
        # Step 5: Add metadata columns (optional)
        if include_metadata:
            cleaned_df = self._add_metadata(cleaned_df, document_source)
        else:
            # Remove validation columns if they exist
            metadata_cols = ['_validation_flags', '_confidence_score', '_needs_review', 
                           '_source_document', '_cleaning_date']
            cleaned_df = cleaned_df.drop(columns=[col for col in metadata_cols if col in cleaned_df.columns])
        
        self.logger.info(f"Data cleaning complete. {len(self.validation_flags)} validation flags generated")
        
        return cleaned_df
    
    def _clean_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize column headers."""
        new_headers = []
        
        for col in df.columns:
            # Convert to string and clean
            header = str(col).strip()
            
            # Apply text fixes
            for pattern, replacement in self.text_fixes.items():
                header = header.replace(pattern, replacement)
            
            # Standardize common education terms
            header = re.sub(r'\bses-?\b', 'session', header, flags=re.IGNORECASE)
            header = re.sub(r'\bundergrad-?\b', 'undergraduate', header, flags=re.IGNORECASE)
            header = re.sub(r'\bfac-?\b', 'faculty', header, flags=re.IGNORECASE)
            header = re.sub(r'\benroll-?\b', 'enrollment', header, flags=re.IGNORECASE)
            
            # Remove extra spaces and special characters
            header = re.sub(r'\s+', ' ', header)
            header = re.sub(r'[^\w\s]', '', header)
            
            # Convert to lowercase for consistency
            header = header.lower().strip()
            
            new_headers.append(header)
        
        df.columns = new_headers
        return df
    
    def _clean_numeric_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean numeric data and fix common OCR errors."""
        for col in df.columns:
            if df[col].dtype == 'object':  # String columns that might contain numbers
                df[col] = df[col].astype(str).apply(self._clean_numeric_value)
        
        return df
    
    def _clean_numeric_value(self, value: str) -> str:
        """Clean individual numeric values."""
        if pd.isna(value) or value in ['', 'nan', 'None']:
            return ''
        
        value = str(value).strip()
        
        # Apply number fixing patterns
        for pattern, replacement in self.number_fixes.items():
            value = re.sub(pattern, replacement, value)
        
        # Remove any remaining non-numeric characters except commas and decimals
        # Keep commas for thousands separators
        cleaned = re.sub(r'[^\d,.]', '', value)
        
        # Handle multiple decimals or commas
        if cleaned.count('.') > 1:
            # Keep only the last decimal point
            parts = cleaned.split('.')
            cleaned = '.'.join(parts[:-1]).replace('.', '') + '.' + parts[-1]
        
        # Validate comma usage (should be every 3 digits from right)
        if ',' in cleaned and re.match(r'^\d{1,3}(,\d{3})*(\.\d+)?$', cleaned):
            return cleaned
        elif ',' in cleaned:
            # Remove improper commas and re-add if needed
            cleaned = cleaned.replace(',', '')
            if cleaned.isdigit() and len(cleaned) > 3:
                # Add commas back properly
                cleaned = f"{int(cleaned):,}"
        
        return cleaned if cleaned else value
    
    def _clean_text_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean text data and fix common OCR errors."""
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).apply(self._clean_text_value)
        
        return df
    
    def _clean_text_value(self, value: str) -> str:
        """Clean individual text values."""
        if pd.isna(value) or value in ['', 'nan', 'None']:
            return ''
        
        value = str(value).strip()
        
        # Apply text fixes
        for pattern, replacement in self.text_fixes.items():
            value = value.replace(pattern, replacement)
        
        # Fix common state name issues
        value = self._fix_state_names(value)
        
        # Remove extra whitespace
        value = re.sub(r'\s+', ' ', value).strip()
        
        return value
    
    def _fix_state_names(self, value: str) -> str:
        """Fix common OCR errors in state names."""
        state_fixes = {
            'Califomia': 'California',
            'Californa': 'California', 
            'New Ycrk': 'New York',
            'New Yark': 'New York',
            'Pennsylvama': 'Pennsylvania',
            'Pennsyivania': 'Pennsylvania',
            'Massachusefts': 'Massachusetts',
            'Massachuselts': 'Massachusetts',
            'Connectlcut': 'Connecticut',
            'Connectlout': 'Connecticut',
        }
        
        for error, correction in state_fixes.items():
            if error.lower() in value.lower():
                value = re.sub(error, correction, value, flags=re.IGNORECASE)
        
        return value
    
    def _validate_data(self, df: pd.DataFrame, document_source: str) -> pd.DataFrame:
        """Validate data quality and flag suspicious values."""
        
        # Add validation columns
        df['_validation_flags'] = ''
        df['_confidence_score'] = 1.0
        
        for idx, row in df.iterrows():
            flags = []
            confidence = 1.0
            
            # Check for missing data in key columns
            if self._is_likely_state_row(row):
                key_cols = [col for col in df.columns if any(term in col.lower() 
                           for term in ['institution', 'faculty', 'enrollment'])]
                
                missing_count = sum(1 for col in key_cols if not str(row[col]).strip() or str(row[col]) == '')
                if missing_count > len(key_cols) * 0.3:  # More than 30% missing
                    flags.append('high_missing_data')
                    confidence *= 0.7
            
            # Validate numeric ranges
            for col in df.columns:
                if col.startswith('_'):  # Skip metadata columns
                    continue
                    
                value = str(row[col]).strip()
                if re.match(r'^[\d,]+$', value):
                    try:
                        num_value = int(value.replace(',', ''))
                        
                        # Check for unreasonable values based on column type
                        if 'institution' in col.lower() and (num_value > 1000 or num_value < 0):
                            flags.append(f'suspicious_institution_count_{col}')
                            confidence *= 0.8
                        elif 'faculty' in col.lower() and (num_value > 100000 or num_value < 0):
                            flags.append(f'suspicious_faculty_count_{col}')
                            confidence *= 0.8
                        elif 'enrollment' in col.lower() and (num_value > 1000000 or num_value < 0):
                            flags.append(f'suspicious_enrollment_{col}')
                            confidence *= 0.8
                        elif 'degree' in col.lower() and (num_value > 100000 or num_value < 0):
                            flags.append(f'suspicious_degree_count_{col}')
                            confidence *= 0.8
                            
                    except ValueError:
                        flags.append(f'invalid_number_{col}')
                        confidence *= 0.5
            
            # Store validation results
            df.at[idx, '_validation_flags'] = '; '.join(flags)
            df.at[idx, '_confidence_score'] = round(confidence, 2)
            
            # Track flags for summary
            if flags:
                self.validation_flags.extend([(document_source, idx, flag) for flag in flags])
        
        return df
    
    def _is_likely_state_row(self, row: pd.Series) -> bool:
        """Check if a row likely represents state-level data."""
        first_col = str(row.iloc[0]).strip()
        # Check if first column looks like a state name
        return (len(first_col) > 2 and 
                first_col[0].isupper() and 
                any(c.islower() for c in first_col) and
                'total' not in first_col.lower())
    
    def _add_metadata(self, df: pd.DataFrame, document_source: str) -> pd.DataFrame:
        """Add metadata columns for tracking."""
        df['_source_document'] = document_source
        df['_cleaning_date'] = datetime.now().isoformat()
        df['_needs_review'] = df['_confidence_score'] < 0.8
        
        return df
    
    def generate_validation_report(self, df: pd.DataFrame, output_path: str):
        """Generate a validation report for research assistants."""
        
        # Summary statistics
        total_rows = len(df)
        flagged_rows = len(df[df['_validation_flags'] != ''])
        needs_review = len(df[df['_needs_review'] == True])
        avg_confidence = df['_confidence_score'].mean()
        
        # Detailed flag breakdown
        flag_summary = {}
        for _, _, flag in self.validation_flags:
            flag_summary[flag] = flag_summary.get(flag, 0) + 1
        
        report = {
            'summary': {
                'total_rows': total_rows,
                'flagged_rows': flagged_rows,
                'needs_review': needs_review,
                'average_confidence': round(avg_confidence, 3),
                'review_percentage': round((needs_review / total_rows) * 100, 1)
            },
            'flag_breakdown': flag_summary,
            'validation_details': self.validation_flags,
            'review_priority': df[df['_needs_review'] == True][['_source_document', '_confidence_score', '_validation_flags']].to_dict('records')
        }
        
        # Save report
        with open(output_path, 'w') as f:
            json.dump(report, f, indent=2)
        
        self.logger.info(f"Validation report saved to {output_path}")
        self.logger.info(f"Summary: {needs_review}/{total_rows} rows need review ({report['summary']['review_percentage']}%)")
        
        return report
    
    def export_for_review(self, df: pd.DataFrame, output_path: str):
        """Export data in format optimized for research assistant review."""
        
        # Create review-friendly version
        review_df = df.copy()
        
        # Reorder columns to put validation info first
        meta_cols = [col for col in review_df.columns if col.startswith('_')]
        data_cols = [col for col in review_df.columns if not col.startswith('_')]
        
        review_df = review_df[['_needs_review', '_confidence_score', '_validation_flags'] + data_cols + 
                             ['_source_document', '_cleaning_date']]
        
        # Sort by confidence score (lowest first - highest priority for review)
        review_df = review_df.sort_values('_confidence_score')
        
        # Export with formatting
        review_df.to_csv(output_path, index=False)
        
        # Also create Excel version with formatting
        excel_path = str(output_path).replace('.csv', '_review.xlsx')
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            review_df.to_excel(writer, sheet_name='Data_Review', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Data_Review']
            
            # Format cells that need review
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            
            for row in range(2, len(review_df) + 2):  # Skip header row
                needs_review = worksheet.cell(row=row, column=1).value
                confidence = worksheet.cell(row=row, column=2).value
                
                if needs_review:
                    for col in range(1, len(review_df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = red_fill
                elif confidence < 0.9:
                    for col in range(1, len(review_df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = yellow_fill
        
        self.logger.info(f"Review files exported: {output_path} and {excel_path}")

def main():
    """Example usage of the data cleaner."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Clean and validate Textract education data')
    parser.add_argument('input_file', help='Path to CSV file from Textract')
    parser.add_argument('--output-dir', '-o', default='test_results', help='Output directory')
    parser.add_argument('--no-metadata', action='store_true', 
                       help='Exclude validation metadata columns from output')
    
    args = parser.parse_args()
    
    # Initialize cleaner
    cleaner = HistoricalDataCleaner()
    
    # Load data
    df = pd.read_csv(args.input_file)
    
    # Clean data
    document_name = Path(args.input_file).stem
    cleaned_df = cleaner.clean_dataframe(df, document_name, include_metadata=not args.no_metadata)
    
    # Generate outputs
    output_dir = Path(args.output_dir)
    output_dir.mkdir(exist_ok=True)
    
    # Save cleaned data
    cleaned_path = output_dir / f"{document_name}_cleaned.csv"
    cleaned_df.to_csv(cleaned_path, index=False)
    
    # Generate validation report only if metadata is included
    if not args.no_metadata:
        report_path = output_dir / f"{document_name}_validation_report.json"
        cleaner.generate_validation_report(cleaned_df, report_path)
        
        # Export for review
        review_path = output_dir / f"{document_name}_for_review.csv"
        cleaner.export_for_review(cleaned_df, review_path)
        
        print(f"✅ Data cleaning complete!")
        print(f"   Cleaned data: {cleaned_path}")
        print(f"   Validation report: {report_path}")
        print(f"   Review files: {review_path}")
    else:
        print(f"✅ Data cleaning complete!")
        print(f"   Cleaned data (no metadata): {cleaned_path}")

if __name__ == "__main__":
    main()